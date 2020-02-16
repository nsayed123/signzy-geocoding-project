#!/usr/bin/python3

import os
import sys
import openpyxl
import requests
import time

#to upload the excel file and test the filepath
def upload_xlsx():
    if len(sys.argv) > 2:
        print('You have specified too many arguments')
        sys.exit()

    if len(sys.argv) < 2:
        print('You need to specify the path to be listed')
        sys.exit()
    input_path = sys.argv[1]
    return input_path

#read the region/address from excel file
def read_xlsx(file_path):

   # Give the location of the file 
    path = file_path
    #print(path)
    # To open the workbook  
    # workbook object is created 
    wb_obj = openpyxl.load_workbook(path) 
  
    # Get workbook active sheet object 
    # from the active attribute 
    sheet_obj = wb_obj.active 
  
    # Cell objects also have row, column,  
    # and coordinate attributes that provide 
    # location information for the cell. 
  
    # Note: The first row or  
    # column integer is 1, not 0.

    #Get the max rows
    max_row = sheet_obj.max_row

    # Cell object is created by using  
    # sheet object's cell() method. 
    cell_obj = sheet_obj.cell(row = 2, column = 1)

    address_list = []
    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        #print(cell_obj.value)
        address_list.append(cell_obj.value)
        #print(list_region)

    return address_list, max_row


#Make an api call to Goole Geocoding API service
def make_api_request(address):
    URL = "https://maps.googleapis.com/maps/api/geocode/json"
    key = "AIzaSyDIP5bqfukMpXJIok6JG_THkCPrZ3Z92Co"
    result_dict = {}
    result_list = []
    for i in address:
        PARAMS = {'address':i, "key":key}
        print(PARAMS)
        r = requests.get(url = URL, params = PARAMS)
        #print(r.url)
        data = r.json()
        #print(data)
        # extracting latitude, longitude and formatted address
        # of the first matching location
        #print(data['results'][0])
        latitude = data['results'][0]['geometry']['location']['lat']
        longitude = data['results'][0]['geometry']['location']['lng']
        formatted_address = data['results'][0]['formatted_address']
        result_dict[i] = {"latitude":latitude, "longitude": longitude, "formatted_address":formatted_address }

        #print("Latitude: {0}\nLongitude: {1}\nFormatted Address: {2}".format(latitude,longitude,formatted_address))
    return  result_dict

#to write the output to excel file
#download the file to the current location
def write_to_excel(address, dict_result, max_row):
    wb = openpyxl.Workbook()
    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active

    # Cell objects also have row, column
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or column integer
    # is 1, not 0. Cell object is created by
    # using sheet object's cell() method.
    sheet.cell(row=1, column=1).value = "Region"

    sheet.cell(row=1, column=2).value = "Latitude"

    sheet.cell(row=1, column=3).value = "Longitude"

    sheet.cell(row=1, column=4).value = "Formatted Address"
    key = [i for i in dict_result.keys()]
    values = [i for i in dict_result.values()]
    print(key)
    print(values)
    for i in range(2, max_row+1):
        j = i - 2
        print(j)
        sheet.cell(row=i, column=1).value = key[j]
        # writing values to cells
        sheet.cell(row=i, column=2). value = values[j]["latitude"]

        sheet.cell(row=i, column=3).value = values[j]["longitude"]
        sheet.cell(row=i, column=4).value = values[j]["formatted_address"]

    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    #a = datetime.now()
    epoch_time = int(time.time())
    filename = str(epoch_time)+".xlsx"
    wb.save(filename)
    print("File downloaded to current location")
    print("Filename: ", filename )


upload = upload_xlsx()
address_list, max_row = read_xlsx(upload)
#print(address_list)

dict_result = make_api_request(address_list)
#print(dict_result)
# latitude, longitude, formatted_address = make_api_request(address)
write_to_excel(address_list ,dict_result,max_row)

#print(latitude, longitude, formatted_address)
